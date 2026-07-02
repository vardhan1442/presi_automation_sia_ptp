#!/usr/bin/env python3
"""
PTP Dashboard — local SSH bridge server.

Provides a REST/SSE API that the browser dashboard uses to:
  1. Test SSH connectivity to a remote host
  2. Pull specific output files (dashboard_data.json + BW sample CSVs)
  3. Trigger verification.py on the remote host and stream its output live

Usage:
    pip install fastapi uvicorn asyncssh
    python tools/server.py
         — or —
    uvicorn tools.server:app --host 127.0.0.1 --port 5000

The server binds to localhost only, so it is not reachable from the network.
"""

import asyncio
import json
import os
import shlex
import traceback
from pathlib import Path, PurePosixPath
from typing import AsyncGenerator, Optional

import asyncssh
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(title="PTP Dashboard SSH Bridge", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*", "null"],  # "null" covers file:// origin in Chrome/Edge
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Constants — files the dashboard needs from each run directory
# ---------------------------------------------------------------------------

DASHBOARD_JSON  = "dashboard_data.json"
IDI_SUBDIR      = "chainsaw_tmp"
DDR_SUBDIR      = "chainsaw_lpddr"
BW_SAMPLE_GLOB  = "overall_bw_samples.txt"   # matches .txt and .txt.gz, not .png.gz

# ---------------------------------------------------------------------------
# Request/response models
# ---------------------------------------------------------------------------

class SSHParams(BaseModel):
    host:        str
    user:        str
    port:        int = 22
    key_content: Optional[str] = None   # raw private key text sent from the browser
    key_path:    Optional[str] = None   # fallback: path on the server's local filesystem
    password:    Optional[str] = None   # password-based auth (used when no key available)

class LoadRequest(SSHParams):
    run_dir: str

    @property
    def run_dir_clean(self) -> str:
        return self.run_dir.rstrip("/")

class RunRequest(SSHParams):
    run_dir:         str
    verification_py: str = "verification.py"
    python_bin:      str = ""   # empty = auto-detect on remote host

    @property
    def run_dir_clean(self) -> str:
        return self.run_dir.rstrip("/")

class BrowseRequest(SSHParams):
    path: str = "~"

    @property
    def path_clean(self) -> str:
        return self.path.rstrip("/") or "/"

# ---------------------------------------------------------------------------
# SSH connection helper
# ---------------------------------------------------------------------------

# Common private-key filenames to try when the specified key is missing
_FALLBACK_KEY_NAMES = ["id_rsa", "id_ed25519", "id_ecdsa", "id_dsa"]


async def _connect(p: SSHParams) -> asyncssh.SSHClientConnection:
    """Open an SSH connection: password → key_content → key_path fallback."""

    # ── Option 0: plain password / keyboard-interactive authentication ────────
    if p.password:
        return await asyncssh.connect(
            host=p.host,
            port=p.port,
            username=p.user,
            password=p.password,
            preferred_auth="password,keyboard-interactive",
            known_hosts=None,
        )

    # ── Option 1: key content sent directly from the browser ──────────────
    if p.key_content:
        try:
            private_key = asyncssh.import_private_key(p.key_content)
        except Exception as exc:
            raise ValueError(f"Could not parse private key: {exc}") from exc

        return await asyncssh.connect(
            host=p.host,
            port=p.port,
            username=p.user,
            client_keys=[private_key],
            agent_path=None,
            known_hosts=None,
        )

    # ── Option 2: path-based fallback (server-local key file) ─────────────
    ssh_dir = Path("~/.ssh").expanduser()
    raw_path = p.key_path or "~/.ssh/id_ed25519"
    requested = Path(raw_path).expanduser().resolve()

    if requested.suffix == ".pub":
        raise ValueError(
            f"'{requested.name}' is a public key file. "
            f"Use the matching private key: {requested.with_suffix('')}"
        )

    key_candidates: list[str] = []
    if requested.exists():
        key_candidates.append(str(requested))
    for name in _FALLBACK_KEY_NAMES:
        candidate = ssh_dir / name
        if candidate.exists() and str(candidate.resolve()) not in key_candidates:
            key_candidates.append(str(candidate.resolve()))

    if not key_candidates:
        raise PermissionError(
            f"No private key found. Upload a key file in the browser, or ensure "
            f"a key exists in {ssh_dir} on the server."
        )

    last_exc: Exception = Exception("No keys to try")
    for key_path in key_candidates:
        try:
            return await asyncssh.connect(
                host=p.host,
                port=p.port,
                username=p.user,
                client_keys=[key_path],
                agent_path=None,
                known_hosts=None,
            )
        except (asyncssh.PermissionDenied, asyncssh.KeyImportError) as exc:
            last_exc = exc
            continue

    raise PermissionError(
        f"Permission denied for user {p.user} on host {p.host}. "
        f"Tried keys: {', '.join(key_candidates)}. "
        f"Make sure the public key is in {p.user}@{p.host}:~/.ssh/authorized_keys."
    )

async def _read_remote(conn: asyncssh.SSHClientConnection, path: str, timeout: int = 20) -> Optional[str]:
    """Read a remote file (plain text OR gzip-compressed) via SSH.

    Uses gzip -dc for .gz files so we never send raw binary data through
    asyncssh's UTF-8 decoder (which would corrupt the SSH connection).
    encoding=None puts asyncssh in binary mode; we decode ourselves with
    errors='replace' so a bad byte never crashes the connection.
    """
    cmd = (
        f"sh -c 'case \"$1\" in "
        f"*.gz) gzip -dc \"$1\" 2>/dev/null;; "
        f"*) cat \"$1\" 2>/dev/null;; "
        f"esac' _ {shlex.quote(path)}"
    )
    result = await conn.run(cmd, check=False, encoding=None, timeout=timeout)
    if not result.stdout:
        return None
    return result.stdout.decode("utf-8", errors="replace")

async def _find_remote(conn: asyncssh.SSHClientConnection, base: str, name_pattern: str) -> Optional[str]:
    """Find the first file under `base` whose name contains `name_pattern`."""
    result = await conn.run(
        f"sh -c 'find \"$1\" -type f -name \"*$2*\" 2>/dev/null | head -1' _ "
        f"{shlex.quote(base)} {shlex.quote(name_pattern)}",
        check=False,
    )
    path = result.stdout.strip()
    return path if path else None

async def _load_backup_cluster(conn: asyncssh.SSHClientConnection, run_dir: str) -> Optional[str]:
    """Load and merge data_collection_backup_01-05.json + backup_global_variables.json.

    Uses a SINGLE ssh round-trip: one shell script cats each file separated by
    a unique delimiter so we can split server-side without 6 separate SSH calls.
    """
    SEP = "__JSON_SEP__"
    # Build a shell snippet that prints each file (if it exists) with a separator
    files = " ".join(
        f"data_collection_backup_{i:02d}.json" for i in range(1, 6)
    ) + " backup_global_variables.json"
    script = (
        f'cd {shlex.quote(run_dir)} 2>/dev/null || exit 0\n'
        f'for f in {files}; do '
        f'[ -f "$f" ] && {{ echo "{SEP}$f"; cat "$f"; echo; }}; '
        f'done'
    )
    try:
        result = await conn.run(
            "sh -s", check=False, input=script, timeout=45
        )
    except Exception as exc:
        print(f"[backup-cluster] read error: {type(exc).__name__}: {exc}")
        return None

    raw = (result.stdout or "").strip()
    if not raw.strip():
        return None

    merged: dict = {}
    any_found = False
    for chunk in raw.split(SEP):
        chunk = chunk.strip()
        if not chunk:
            continue
        # First line is the filename, rest is JSON content
        nl = chunk.find("\n")
        if nl < 0:
            continue
        fname = chunk[:nl].strip()
        content = chunk[nl:].strip()
        if not content:
            continue
        try:
            data = json.loads(content)
            any_found = True
            if "backup_global_variables" in fname:
                for key in ("rangeLoops", "typeModel", "soc_cdie", "workload",
                            "endtimeStamp", "StarttimeStamp", "cores_found"):
                    if key in data:
                        merged.setdefault(key, data[key])
            else:
                merged.update(data)
        except Exception:
            pass

    print(f"[backup-cluster] any_found={any_found}, keys={list(merged.keys())[:8]}")
    return json.dumps(merged) if any_found else None

async def _run_extractor(conn: asyncssh.SSHClientConnection, run_dir: str) -> Optional[str]:
    """Pipe tools/extract_ptp_data.py to the remote Python interpreter.

    The extractor runs entirely on the remote host (same Python that's in PATH
    there) and returns a single JSON line to stdout.  No file-copying needed —
    the script content is sent via stdin.
    """
    extractor_path = Path(__file__).parent / "extract_ptp_data.py"
    if not extractor_path.exists():
        print(f"[extractor] {extractor_path} not found — skipping")
        return None

    script = extractor_path.read_text(encoding="utf-8")

    # Auto-detect Python: prefer python3.11.1, then the Intel NFS full path, then python3
    cmd = (
        f"sh -c '"
        f"_py=$(command -v python3.11.1 2>/dev/null || "
        f"command -v /nfs/site/itools/em64t_SLES12SP5/pkgs/python3/3.11.1/bin/python3.11.1 2>/dev/null || "
        f"command -v python3 2>/dev/null || echo python3) && "
        f"cd {shlex.quote(run_dir)} && \"$_py\" - -r {shlex.quote(run_dir)}"
        f"'"
    )
    print(f"[extractor] running on remote: {cmd[:120]}…")
    try:
        result = await conn.run(cmd, check=False, input=script, timeout=25)
        stdout = (result.stdout or "").strip()
        if result.returncode != 0:
            print(f"[extractor] rc={result.returncode}  stderr={result.stderr[:300]!r}")
        if stdout and stdout.startswith("{"):
            print(f"[extractor] JSON received ({len(stdout)} bytes)")
            return stdout
        print(f"[extractor] unexpected stdout: {stdout[:200]!r}")
        return None
    except Exception as exc:
        print(f"[extractor] error: {type(exc).__name__}: {exc}")
        return None

async def _get_dir_times(
    conn: asyncssh.SSHClientConnection, run_dir: str
) -> tuple[dict, str]:
    """
    Collect modification + change timestamps for the run directory and key
    sub-entries, plus the full ls -la listing.

    Returns:
        dir_times : dict  — { basename: {"create": ts, "modify": ts} }
        ls_output : str   — raw ls -la listing of the run directory
    """
    targets = [
        run_dir,
        f"{run_dir}/chainsaw_tmp",
        f"{run_dir}/chainsaw_lpddr",
        f"{run_dir}/dashboard_data.json",
    ]
    # Pass paths as positional "$@" so the format string can safely use double
    # quotes inside the single-quoted sh -c script — no nested quote conflicts.
    target_args = " ".join(shlex.quote(t) for t in targets)

    # Single stat call: name | modification-time | change-time (ctime as
    # creation proxy; Linux rarely supports true birth time via %w).
    stat_cmd = (
        f"sh -c 'stat -c \"%n|%y|%z\" \"$@\" 2>/dev/null' -- {target_args}"
    )
    stat_result = await conn.run(stat_cmd, check=False)
    print(f"[dir-times] stat stdout: {stat_result.stdout!r}")
    print(f"[dir-times] stat stderr: {stat_result.stderr!r}")

    dir_times: dict = {}
    for line in (stat_result.stdout or "").splitlines():
        line = line.strip()
        parts = line.split("|", 2)
        if len(parts) < 2:
            continue
        name   = parts[0].rstrip("/").strip()
        mtime  = parts[1].strip()[:19]
        ctime  = parts[2].strip()[:19] if len(parts) > 2 else ""
        key    = name.split("/")[-1] if "/" in name else name
        if not key or not mtime:
            continue
        dir_times[key] = {
            "modify": mtime,
            "create": ctime if ctime and ctime != "-" else mtime,
        }

    # Full ls -la listing (with long-iso timestamps where available)
    ls_cmd = (
        f"sh -c 'ls -la --time-style=long-iso \"$1\" 2>/dev/null"
        f" || ls -la \"$1\" 2>/dev/null' -- {shlex.quote(run_dir)}"
    )
    ls_result = await conn.run(ls_cmd, check=False)
    ls_output = (ls_result.stdout or "").strip()

    print(f"[dir-times] dir_times={dir_times}")
    print(f"[dir-times] ls_lines={len(ls_output.splitlines())}")
    return dir_times, ls_output


# ---------------------------------------------------------------------------
# Shell-only loop info fallback (no Python needed on remote)
# ---------------------------------------------------------------------------

async def _get_loop_info_shell(
    conn: asyncssh.SSHClientConnection, run_dir: str
) -> Optional[str]:
    """Extract minimal loop metadata via a single compound shell script.

    Runs when the Python extractor fails/times-out and no backup JSON files
    exist.  One SSH round-trip, 45 s total timeout.
    """
    script = (
        'cd "$1" 2>/dev/null || exit 0\n'
        # Loops / Lsz from LST file
        "LOOPS=$(zgrep 'mov r13, 0x' *lst* 2>/dev/null | tail -1 | awk '{print $11}')\n"
        "LSZ=$(zgrep   'mov r10, 0x' *lst* 2>/dev/null | tail -1 | awk '{print $11}')\n"
        '[ -n "$LOOPS" ] && echo "Loops=$LOOPS"\n'
        '[ -n "$LSZ" ]   && echo "Lsz=$LSZ"\n'
        # IDI test window
        "TS=$(zgrep -i '[0]*0[234][0-9a-f]\\{8\\} ' idi_bridge.log* 2>/dev/null"
        " | grep -i 'c2u_req' | head -1)\n"
        "TE=$(zgrep -i '[0]*0[234][0-9a-f]\\{8\\} ' idi*.log* 2>/dev/null"
        " | grep -i 'c2u_req' | grep -iv 'wb.fto.' | tail -1)\n"
        '[ -n "$TS" ] && echo "testStart=$TS"\n'
        '[ -n "$TE" ] && echo "testEnd=$TE"\n'
        # DDR test window
        "DS=$(zgrep -i 'MEM_READ' trackers_socn/cmi_bw/cmi_jem_tracker.log* 2>/dev/null | head -1)\n"
        "DE=$(zgrep -i 'MEM_READ' trackers_socn/cmi_bw/cmi_jem_tracker.log* 2>/dev/null | tail -1)\n"
        '[ -n "$DS" ] && echo "testStartDDR=$DS"\n'
        '[ -n "$DE" ] && echo "testEndDDR=$DE"\n'
        # Loop boundary addresses from LST
        "LOOP_LIP=$(zgrep -i 00abcde do*.lst* 2>/dev/null"
        " | awk -F ':' '{print $1}' | awk -F 'x' '{print $2}' | head -1)\n"
        "START_LIP=$(zgrep -i 'mov.*xmm.*[[rsdcxi]]' do*.lst* 2>/dev/null"
        " | head -1 | awk -F ':' '{print $1}' | awk -F 'x' '{print $2}')\n"
        '[ -z "$START_LIP" ] && START_LIP=$(zgrep -i '
        "'mov r8, rax' do*.lst* 2>/dev/null"
        " | head -1 | awk -F ':' '{print $1}' | awk -F 'x' '{print $2}')\n"
        '[ -n "$LOOP_LIP" ]  && echo "loop_lip=$LOOP_LIP"\n'
        '[ -n "$START_LIP" ] && echo "start_lip_addr=$START_LIP"\n'
        # LIP tracker timestamps + full log lines
        'if [ -n "$LOOP_LIP" ] && [ -n "$START_LIP" ]; then\n'
        '  END_T=$(zgrep -i "$LOOP_LIP"  lip*tracker*.log* 2>/dev/null'
        " | awk -F ':' '{print $2}' | awk -F '|' '{print $1}')\n"
        '  STA_T=$(zgrep -i "$START_LIP" lip*tracker*.log* 2>/dev/null'
        " | head -1 | awk -F ':' '{print $2}' | awk -F '|' '{print $1}')\n"
        '  [ -n "$END_T" ] && printf "end_times=%s\\n" "$END_T"\n'
        '  [ -n "$STA_T" ] && printf "start_times=%s\\n" "$STA_T"\n'
        '  LIP_START_LINE=$(zgrep -i "$START_LIP" lip*tracker*.log* 2>/dev/null | head -1)\n'
        '  LIP_END_LINE=$(zgrep -i "$LOOP_LIP"   lip*tracker*.log* 2>/dev/null | tail -1)\n'
        '  [ -n "$LIP_START_LINE" ] && echo "lip_start_line=$LIP_START_LINE"\n'
        '  [ -n "$LIP_END_LINE" ]   && echo "lip_end_line=$LIP_END_LINE"\n'
        'fi\n'
    )
    try:
        result = await conn.run(
            f"sh -s -- {shlex.quote(run_dir)}",
            check=False, input=script, timeout=90
        )
    except Exception as exc:
        print(f"[shell-fallback] error: {type(exc).__name__}: {exc}")
        return None

    stdout = (result.stdout or "").strip()
    if not stdout:
        print("[shell-fallback] no output")
        return None

    data: dict = {}
    kv: dict = {}
    for line in stdout.splitlines():
        if "=" in line:
            k, _, v = line.partition("=")
            kv[k.strip()] = v.strip()

    if kv.get("Loops"): data["Loops"] = kv["Loops"]
    if kv.get("Lsz"):   data["Lsz"]   = kv["Lsz"]
    for key in ("testStart", "testEnd", "testStartDDR", "testEndDDR"):
        if kv.get(key): data[key] = kv[key]

    # LIP test boundaries: full log lines for start and end
    if kv.get("lip_start_line") or kv.get("lip_end_line"):
        data["start_lip"] = [
            "### Start ###",
            kv.get("lip_start_line") or "none",
            "### End ###",
            kv.get("lip_end_line")   or "none",
        ]

    # Reconstruct rangeLoops from LIP tracker timestamps
    if kv.get("end_times") and kv.get("start_times"):
        end_times   = [t.strip() for t in kv["end_times"].splitlines()   if t.strip().lstrip("-").isdigit()]
        start_times = [t.strip() for t in kv["start_times"].splitlines() if t.strip().lstrip("-").isdigit()]
        if end_times and start_times:
            num_cores = len(start_times)
            n_loops   = len(end_times) // max(num_cores, 1)
            range_loops = []
            for cc in range(n_loops):
                lo     = int(start_times[0])
                hi_idx = num_cores * cc + (num_cores - 1)
                hi     = int(end_times[hi_idx]) if hi_idx < len(end_times) else int(end_times[-1])
                range_loops.append([str(lo), str(hi)])
            if range_loops:
                data["rangeLoops"] = range_loops

    if not data:
        print("[shell-fallback] no data extracted")
        return None
    print(f"[shell-fallback] keys: {list(data.keys())}")
    return json.dumps(data)


# ---------------------------------------------------------------------------
# Dedicated opcode / DDR extraction endpoint
# ---------------------------------------------------------------------------

async def _read_xlsx_opcodes(conn: asyncssh.SSHClientConnection, run_dir: str) -> Optional[dict]:
    """Download the verification.py-generated xlsx and parse opcode + DDR data.

    Uses asyncssh SFTP for the download, openpyxl read_only mode for speed.
    Rejects files larger than 80 MB to avoid memory pressure.
    """
    import io
    try:
        import openpyxl
    except ImportError:
        print("[xlsx] openpyxl not installed — skipping xlsx parse")
        return None

    # Locate xlsx
    r = await conn.run(
        f"ls {shlex.quote(run_dir)}/*.xlsx 2>/dev/null | head -1",
        check=False, timeout=10
    )
    xlsx_path = (r.stdout or "").strip()
    if not xlsx_path:
        return None

    # Check file size (reject > 80 MB)
    sz_r = await conn.run(
        f"stat -c '%s' {shlex.quote(xlsx_path)} 2>/dev/null",
        check=False, timeout=10
    )
    size_bytes = int((sz_r.stdout or "0").strip() or "0")
    if size_bytes > 80 * 1024 * 1024:
        print(f"[xlsx] {xlsx_path} too large ({size_bytes // 1024 // 1024} MB) — skipping")
        return None
    print(f"[xlsx] downloading {xlsx_path} ({size_bytes // 1024} KB)")

    # Download via SFTP
    buf = io.BytesIO()
    async with conn.start_sftp_client() as sftp:
        await sftp.getfo(xlsx_path, buf)
    buf.seek(0)

    wb  = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    ws  = wb.active
    out: dict = {}

    in_opcodes = False
    in_ddr     = False
    opcodes: list = []
    ddr_s: list   = []

    # Scan the active sheet row-by-row
    for row in ws.iter_rows(values_only=True):
        # Column B = index 1 (0-based), Column C = index 2
        col_b = str(row[1] or "").strip() if len(row) > 1 else ""
        col_c = str(row[2] or "").strip() if len(row) > 2 else ""

        # ── Section markers ────────────────────────────────────────────────
        if "Opcodes Seen" in col_b and "IDI" in col_b:
            in_opcodes, in_ddr = True, False
            continue
        if "DDR" in col_b and any(k in col_b for k in ("Read", "Write", "info", "lpddr")):
            in_ddr, in_opcodes = True, False
            continue
        # New major section resets state (col B non-empty label ≥ 8 chars)
        if col_b and len(col_b) >= 8 and col_b not in ("IA|AT", "IA,AT/Loop"):
            if in_opcodes or in_ddr:
                in_opcodes, in_ddr = False, False

        # ── Opcode rows ────────────────────────────────────────────────────
        if in_opcodes and col_c:
            # tabledinamic writes opcode name in col C (index 2)
            # subsequent columns may have count / type / mode
            row_vals = [str(v or "").strip() for v in row[2:7]]
            non_empty = [v for v in row_vals if v]
            if non_empty:
                # Treat first non-empty cell as opcode name, rest as metadata
                opcodes.append(row_vals)

        # ── DDR opcode breakdown ──────────────────────────────────────────
        if in_ddr and col_c and col_b in ("", "IA|AT"):
            row_vals = [str(v or "").strip() for v in row[2:5]]
            if row_vals[0]:
                ddr_s.append(", ".join(v for v in row_vals if v))

    # Consolidate
    if opcodes:
        # Re-shape: [opcode_name, count, type, mode] — col order from tabledinamic
        reshaped = []
        for r in opcodes:
            name = r[0]; rest = r[1:]
            reshaped.append([rest[0] if rest else "", name, rest[1] if len(rest) > 1 else "", ""])
        out["opcodes"] = reshaped
        out["_source"] = "xlsx"
        print(f"[xlsx] extracted {len(reshaped)} opcodes")
    if ddr_s:
        out["DDR_infoS"] = ddr_s

    # Also try to pull testStart / testEnd / Loops / Lsz from the sheet
    for row in ws.iter_rows(values_only=True):
        col_b = str(row[1] or "").strip() if len(row) > 1 else ""
        col_c = str(row[2] or "").strip() if len(row) > 2 else ""
        if col_b == "Test Start IDI:" and col_c:
            out.setdefault("testStart", col_c)
        elif col_b == "Test End:" and col_c:
            out.setdefault("testEnd", col_c)
        elif col_b == "Loops" and col_c:
            out.setdefault("Loops", col_c)
        elif col_b == "Lsz:" and col_c:
            out.setdefault("Lsz", col_c)

    wb.close()
    return out or None

async def _run_opcode_script(conn: asyncssh.SSHClientConnection, run_dir: str) -> dict:
    """Run a focused shell script to extract opcode + DDR stats.

    Uses head to limit reads from large tracker files so the pipeline exits
    early.  Encodes multi-value results as pipe-separated (~-delimited) lines.
    """
    script = (
        'cd "$1" 2>/dev/null || exit 0\n'
        # Report which tracker files exist (fast ls)
        "IDI_FILES=$(ls idi*.log* 2>/dev/null | head -5 | tr '\\n' '|')\n"
        "MC_FILES=$(ls MC*.log*  2>/dev/null | head -5 | tr '\\n' '|')\n"
        "LP_FILES=$(ls lpddr5_xtor*tracker* 2>/dev/null | head -5 | tr '\\n' '|')\n"
        '[ -n "$IDI_FILES" ] && echo "idi_files=$IDI_FILES"\n'
        '[ -n "$MC_FILES"  ] && echo "mc_files=$MC_FILES"\n'
        '[ -n "$LP_FILES"  ] && echo "lp_files=$LP_FILES"\n'
        # IDI opcodes: read first 100K lines then aggregate (head kills pipeline early)
        "OPCODES=$(zcat idi*.log* 2>/dev/null | head -100000"
        " | awk -F '|' 'NR>1 && $7!=\"\" && $7!=\"-\" && $7!=\"Unit\""
        " {gsub(/ /,\"\",$7); print $7}'"
        " | sort | uniq -c | sort -rn | head -30"
        " | awk '{printf \"%s~%s|\", $1, $2}')\n"
        '[ -n "$OPCODES" ] && echo "opcodes_raw=$OPCODES"\n'
        # DDR read/write totals from MC logs
        "DDR_RD=$(zgrep -ic 'Rd \\|_RD' MC*.log* 2>/dev/null"
        " | awk -F: '{s+=$2} END {print s+0}')\n"
        "DDR_WR=$(zgrep -ic 'Wr \\|_WR' MC*.log* 2>/dev/null"
        " | awk -F: '{s+=$2} END {print s+0}')\n"
        '[ "${DDR_RD:-0}" != "0" ] && echo "DDR_info=$DDR_RD"\n'
        '[ "${DDR_WR:-0}" != "0" ] && echo "DDR_infoW=$DDR_WR"\n'
        # DDR opcode breakdown from lpddr5 tracker
        "DDR_S=$(zgrep -i '|' lpddr5_xtor*tracker* 2>/dev/null"
        " | awk -F '|' 'NF>5 {gsub(/ /,\"\",$6); if($6!=\"\") print $6}'"
        " | sort | uniq -c | sort -rn | head -20"
        " | awk '{printf \"%s~%s|\", $1, $2}')\n"
        '[ -n "$DDR_S" ] && echo "ddr_ops=$DDR_S"\n'
        # SOC CFI opcode breakdown
        "CFI_OPS=$(zcat SOC_CFI_trk*.log* 2>/dev/null | head -100000"
        " | awk -F '|' 'NF>6 {gsub(/ /,\"\",$2); gsub(/ /,\"\",$6); gsub(/ /,\"\",$7);"
        " if($2!=\"\") printf \"%s~%s~%s|\", $2, $6, $7}'"
        " | tr '|' '\\n' | sort | uniq -c | sort -rn | head -30"
        " | awk '{printf \"%s~%s|\", $1, $2}')\n"
        '[ -n "$CFI_OPS" ] && echo "cfi_ops=$CFI_OPS"\n'
    )

    result = await conn.run(
        f"sh -s -- {shlex.quote(run_dir)}",
        check=False, input=script, timeout=120
    )
    stdout = (result.stdout or "").strip()
    print(f"[opcodes] script stdout length: {len(stdout)}")

    out: dict = {}
    kv: dict = {}
    for line in stdout.splitlines():
        if "=" in line:
            k, _, v = line.partition("=")
            kv[k.strip()] = v.strip()

    # Report detected files
    for fk in ("idi_files", "mc_files", "lp_files"):
        if kv.get(fk):
            print(f"[opcodes] {fk}: {kv[fk]}")

    # IDI opcodes: "count~opcode|count~opcode|..."
    if kv.get("opcodes_raw"):
        opcodes = []
        for token in kv["opcodes_raw"].split("|"):
            token = token.strip()
            if "~" in token:
                cnt, op = token.split("~", 1)
                opcodes.append([cnt.strip(), op.strip(), "", ""])
        if opcodes:
            out["opcodes"] = opcodes
            print(f"[opcodes] extracted {len(opcodes)} opcode entries")

    # DDR counts
    if kv.get("DDR_info"):  out["DDR_info"]  = kv["DDR_info"]
    if kv.get("DDR_infoW"): out["DDR_infoW"] = kv["DDR_infoW"]

    # DDR opcode breakdown
    if kv.get("ddr_ops"):
        ddr_s = []
        for token in kv["ddr_ops"].split("|"):
            token = token.strip()
            if "~" in token:
                cnt, op = token.split("~", 1)
                ddr_s.append(f"{cnt.strip()}, {op.strip()}")
        if ddr_s:
            out["DDR_infoS"] = ddr_s

    # SOC CFI ops
    if kv.get("cfi_ops"):
        cfi = []
        for token in kv["cfi_ops"].split("|"):
            token = token.strip()
            if "~" in token:
                parts = token.split("~")
                cfi.append(", ".join(p.strip() for p in parts))
        if cfi:
            out["opcodes_soc_cfi"] = cfi

    if not kv.get("idi_files") and not kv.get("mc_files"):
        out["_warning"] = "No IDI or MC tracker files found in run directory"

    return out


@app.post("/api/load-opcodes")
async def load_opcodes(req: LoadRequest):
    """Load opcode + DDR stats with a hard 150 s deadline.

    Priority:
      1. backup cluster JSON  (data_collection_backup_*.json) — produced by verification.py
      2. xlsx file            (verification.py Excel output) — downloaded via SFTP
      3. dashboard_data.json  (legacy single-file output)
      4. Shell-based tracker grep (slow, sampled first 100K lines)
    """
    import asyncio
    run_dir = req.run_dir_clean
    print(f"[opcodes] request received for run_dir={run_dir}")

    async def _do_load():
        async with await _connect(req) as conn:

            # ── Step 1: try backup cluster JSON ─────────────────────────────
            print("[opcodes] checking backup cluster JSON ...")
            backup_text = await _load_backup_cluster(conn, run_dir)
            if backup_text:
                try:
                    bd = json.loads(backup_text)
                    result: dict = {}
                    for key in ("opcodes", "DDR_info", "DDR_infoW", "DDR_infoS",
                                "opcodes_soc_cfi", "OpcodeLoop", "OpcodeLoopCFI"):
                        if key in bd and bd[key]:
                            result[key] = bd[key]
                    if result:
                        result["_source"] = "backup_json"
                        print(f"[opcodes] loaded from backup JSON — keys: {list(result.keys())}")
                        return result
                except Exception as exc:
                    print(f"[opcodes] backup JSON parse error: {exc}")

            # ── Step 2: try xlsx (verification.py Excel output) ──────────────
            print("[opcodes] checking xlsx ...")
            xlsx_data = await _read_xlsx_opcodes(conn, run_dir)
            if xlsx_data:
                return xlsx_data

            # ── Step 3: try dashboard_data.json (legacy) ─────────────────────
            print("[opcodes] checking dashboard_data.json ...")
            dash_text = await _read_remote(conn, f"{run_dir}/dashboard_data.json")
            if dash_text:
                try:
                    dd = json.loads(dash_text)
                    result = {}
                    for key in ("opcodes", "DDR_info", "DDR_infoW", "DDR_infoS",
                                "opcodes_soc_cfi", "OpcodeLoop"):
                        if key in dd and dd[key]:
                            result[key] = dd[key]
                    if result:
                        result["_source"] = "dashboard_json"
                        print(f"[opcodes] loaded from dashboard_data.json — keys: {list(result.keys())}")
                        return result
                except Exception:
                    pass

            # ── Step 4: check if xlsx exists (already tried above, just log) ──
            xlsx_res = await conn.run(
                f"ls {shlex.quote(run_dir)}/*.xlsx 2>/dev/null | head -1",
                check=False, timeout=10
            )
            xlsx_found = (xlsx_res.stdout or "").strip()
            if xlsx_found:
                print(f"[opcodes] xlsx found ({xlsx_found}) but no backup JSON — falling back to shell extraction")

            # ── Step 5: shell-based tracker grep (sampled) ───────────────────
            print("[opcodes] falling back to shell tracker extraction ...")
            result = await _run_opcode_script(conn, run_dir)
            if not result.get("_warning"):
                result["_source"] = "tracker_sample"
            return result

    try:
        result = await asyncio.wait_for(_do_load(), timeout=150)
        return JSONResponse(result)
    except asyncio.TimeoutError:
        return JSONResponse({"_warning": "Opcode load timed out after 150 s. Run Verification Script to generate backup JSON."})
    except Exception as exc:
        return JSONResponse({"_warning": f"Opcode load failed: {type(exc).__name__}: {exc}"})


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get("/")
def root():
    return {"service": "PTP Dashboard SSH Bridge", "status": "running"}


@app.post("/api/connect-test")
async def connect_test(p: SSHParams):
    """Verify that SSH credentials are valid and the host is reachable."""
    try:
        async with await _connect(p) as conn:
            res = await conn.run("hostname && whoami", check=True)
            lines = res.stdout.strip().splitlines()
            return {
                "ok":       True,
                "hostname": lines[0] if lines else "",
                "user":     lines[1] if len(lines) > 1 else p.user,
            }
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


class ExecRequest(SSHParams):
    command: str
    timeout: int = 30

@app.post("/api/exec-command")
async def exec_command(req: ExecRequest):
    """Stream command output as SSE — same pattern as run-verification so TCP stays alive."""
    async def sse_stream() -> AsyncGenerator[str, None]:
        def event(payload: dict) -> str:
            return f"data: {json.dumps(payload)}\n\n"

        print(f"[exec-command] host={req.host} user={req.user} cmd={req.command[:80]!r}")
        try:
            async with await _connect(req) as conn:
                print(f"[exec-command] connected, creating process")
                # Use tcsh (remote shell) so source, alias, setenv etc. work correctly.
                # Commands are piped via stdin to avoid quoting/escaping issues.
                async with conn.create_process(
                    "tcsh -s", encoding="utf-8",
                    stdin=asyncssh.PIPE, stderr=asyncssh.PIPE
                ) as proc:
                    proc.stdin.write(req.command + "\n")
                    proc.stdin.write_eof()
                    async for line in proc.stdout:
                        yield event({"type": "stdout", "line": line.rstrip("\n")})
                    async for line in proc.stderr:
                        yield event({"type": "stderr", "line": line.rstrip("\n")})
                    es = proc.exit_status
                    rc = es.returncode if hasattr(es, "returncode") else (es if isinstance(es, int) else 0)
                print(f"[exec-command] done, exit={rc}")
                yield event({"type": "done", "rc": rc})
        except Exception as exc:
            print(f"[exec-command] error: {type(exc).__name__}: {exc}")
            yield event({"type": "error", "message": str(exc)})

    return StreamingResponse(
        sse_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/list-dir")
async def list_dir(req: BrowseRequest):
    """
    List directory contents on the remote host.
    Returns the resolved absolute path and a list of entries (name, is_dir).
    """
    try:
        async with await _connect(req) as conn:
            raw = req.path_clean

            # Build a POSIX sh command; pass path as $1 to avoid quoting/expansion
            # issues.  We wrap in 'sh -c ...' so the command runs in POSIX sh even
            # when the user's login shell is tcsh/csh (which doesn't support &&,
            # 2>/dev/null, or $(...) substitution).
            if raw in ("~", ""):
                # bare 'cd' goes to $HOME in every POSIX shell
                cmd = "sh -c 'cd && pwd && ls -1p . 2>/dev/null'"
            elif raw.startswith("~/"):
                # pass the sub-path as $1 so tilde expands correctly in sh
                sub_arg = shlex.quote(raw[2:])
                cmd = f'''sh -c 'cd ~/"$1" && pwd && ls -1p . 2>/dev/null' _ {sub_arg}'''
            else:
                path_arg = shlex.quote(raw)
                cmd = f'''sh -c 'cd "$1" && pwd && ls -1p . 2>/dev/null' _ {path_arg}'''

            result = await conn.run(cmd, check=False)
            print(f"[list-dir] cmd={cmd!r}")
            print(f"[list-dir] rc={result.returncode!r}  stdout={result.stdout[:120]!r}  stderr={result.stderr[:120]!r}")
            if result.returncode not in (0, None) or not result.stdout.strip():
                err = (result.stderr or "").strip()
                raise HTTPException(
                    status_code=404,
                    detail=f"Cannot access directory: {req.path}" + (f" — {err}" if err else ""),
                )
            lines = result.stdout.strip().splitlines()
            abs_path = lines[0] if lines else req.path
            entries = []
            for line in lines[1:]:
                if not line or line in (".", ".."):
                    continue
                is_dir = line.endswith("/")
                entries.append({"name": line.rstrip("/"), "is_dir": is_dir})
            return {"path": abs_path, "entries": entries}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/load-files")
async def load_files(req: LoadRequest):
    """
    Pull files from a remote run directory.
    dashboard_data.json is optional — if absent, the browser renders
    whatever graph data is available.
    """
    try:
        async with await _connect(req) as conn:
            run_dir = req.run_dir_clean
            print(f"[load-files] run_dir={run_dir!r}")

            # Verify the directory is accessible first
            check = await conn.run(
                f"sh -c 'test -d \"$1\" && echo ok' _ {shlex.quote(run_dir)}", check=False
            )
            print(f"[load-files] test-d rc={check.returncode!r} stdout={check.stdout.strip()!r} stderr={check.stderr.strip()!r}")
            if check.stdout.strip() != "ok":
                raise HTTPException(
                    status_code=404,
                    detail=f"Directory not found or not accessible: {run_dir}",
                )

            # 1. Run the lightweight extractor (extract_ptp_data.py) on the remote.
            #    This pipes the local script to the remote Python via stdin — no
            #    file copying needed.  Falls back to cached JSON files if it fails.
            print(f"[load-files] running extractor …")
            json_text = await _run_extractor(conn, run_dir)

            # 2. If extractor produced nothing, fall back to backup cluster files
            #    written by the full verification.py (cluster_n_save).
            if not json_text:
                print(f"[load-files] extractor failed — trying backup cluster files …")
                json_text = await _load_backup_cluster(conn, run_dir)

            # 3. Last resort: dashboard_data.json (legacy)
            if not json_text:
                print(f"[load-files] trying dashboard_data.json …")
                json_text = await _read_remote(conn, f"{run_dir}/{DASHBOARD_JSON}")
            if not json_text:
                found = await _find_remote(conn, run_dir, DASHBOARD_JSON)
                json_text = await _read_remote(conn, found) if found else None

            # 4. Shell-only minimal fallback: extract Loops/Lsz/testStart/testEnd
            #    without Python — runs even when the extractor times out or fails.
            if not json_text:
                print(f"[load-files] running shell-only loop info fallback …")
                json_text = await _get_loop_info_shell(conn, run_dir)

            print(f"[load-files] json_text={'ok' if json_text else 'None'}")

            # 4. IDI BW sample — search inside chainsaw_tmp/
            print(f"[load-files] searching IDI …")
            idi_path = await _find_remote(conn, f"{run_dir}/{IDI_SUBDIR}", BW_SAMPLE_GLOB)
            idi_text = await _read_remote(conn, idi_path) if idi_path else None
            print(f"[load-files] idi_path={idi_path!r}")

            # 5. DDR BW sample — search inside chainsaw_lpddr/
            print(f"[load-files] searching DDR …")
            ddr_path = await _find_remote(conn, f"{run_dir}/{DDR_SUBDIR}", BW_SAMPLE_GLOB)
            ddr_text = await _read_remote(conn, ddr_path) if ddr_path else None
            print(f"[load-files] ddr_path={ddr_path!r}")

            # 6. Directory timestamps via stat + ls -l
            print(f"[load-files] fetching directory timestamps …")
            dir_times, ls_output = await _get_dir_times(conn, run_dir)

            return {
                "json_text": json_text,
                "idi_text":  idi_text,
                "ddr_text":  ddr_text,
                "idi_file":  PurePosixPath(idi_path).name if idi_path else None,
                "ddr_file":  PurePosixPath(ddr_path).name if ddr_path else None,
                "dir_times": dir_times,
                "ls_output": ls_output,
            }

    except HTTPException:
        raise
    except Exception as exc:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(exc))


class ReadFileRequest(SSHParams):
    path: str


@app.post("/api/read-file")
async def read_remote_file(req: ReadFileRequest):
    """Read a single file from the remote host and return its text content."""
    try:
        async with await _connect(req) as conn:
            result = await conn.run(
                f"sh -c 'cat \"$1\"' _ {shlex.quote(req.path)}",
                check=False,
            )
            if result.returncode != 0:
                err = (result.stderr or "").strip()
                raise HTTPException(
                    status_code=404,
                    detail=f"Cannot read file: {req.path}" + (f" — {err}" if err else ""),
                )
            return {"path": req.path, "content": result.stdout}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/run-verification")
async def run_verification(req: RunRequest):
    """
    Run verification.py on the remote host and stream its stdout/stderr back
    to the browser as Server-Sent Events (SSE).

    After the process exits (rc == 0), the browser should call /api/load-files
    to pull the freshly generated dashboard_data.json.
    """
    async def sse_stream() -> AsyncGenerator[str, None]:
        def event(payload: dict) -> str:
            return f"data: {json.dumps(payload)}\n\n"

        try:
            async with await _connect(req) as conn:
                run_dir = req.run_dir_clean
                vpy = req.verification_py
                if not vpy.startswith("/"):
                    vpy = f"{run_dir}/{vpy}"
                python_bin = req.python_bin.strip()

                # Auto-detect Python on remote: try the requested bin first (if provided),
                # then the known Intel NFS path for Python 3.11.1, then python3 / python.
                # All wrapped in sh -c so it works even on tcsh login shells.
                cmd = (
                    f"sh -c '_py=$([ -n \"$3\" ] && command -v \"$3\" 2>/dev/null || "
                    f"command -v python3.11.1 2>/dev/null || "
                    f"command -v /nfs/site/itools/em64t_SLES12SP5/pkgs/python3/3.11.1/bin/python3.11.1 2>/dev/null || "
                    f"command -v python3 2>/dev/null || echo python3.11.1) && "
                    f"cd \"$1\" && \"$_py\" \"$2\" -r \"$1\" 2>&1' _ "
                    f"{shlex.quote(run_dir)} {shlex.quote(vpy)} {shlex.quote(python_bin)}"
                )
                display_cmd = f"{python_bin or '(auto-detect)'} {vpy} -r {run_dir}"
                yield event({"type": "start", "cmd": display_cmd})

                async with conn.create_process(cmd) as proc:
                    async for line in proc.stdout:
                        yield event({"type": "stdout", "line": line.rstrip()})

                rc = await proc.wait()
                yield event({"type": "done", "rc": rc})

        except Exception as exc:
            yield event({"type": "error", "message": str(exc)})

    return StreamingResponse(
        sse_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control":    "no-cache",
            "X-Accel-Buffering": "no",   # disable nginx buffering if proxied
        },
    )


# ---------------------------------------------------------------------------
# Static file serving — project root served at /
# API routes defined above always take priority over static files.
# Open the dashboard at:  http://127.0.0.1:5000/analysis.html
# ---------------------------------------------------------------------------

_project_root = str(Path(__file__).parent.parent)
app.mount("/", StaticFiles(directory=_project_root, html=False), name="static")

# ---------------------------------------------------------------------------
# Dev entry-point
# ---------------------------------------------------------------------------

def _free_port(port: int) -> None:
    """If something is already listening on `port`, kill it (Windows + Linux)."""
    import socket, subprocess, time, sys
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        if s.connect_ex(("127.0.0.1", port)) != 0:
            return  # port is free
    print(f"[server] port {port} is in use — attempting to free it…")
    try:
        if sys.platform == "win32":
            out = subprocess.check_output(["netstat", "-ano"], text=True)
            for line in out.splitlines():
                if f":{port}" in line and "LISTENING" in line:
                    pid = line.strip().split()[-1]
                    subprocess.call(["taskkill", "/pid", pid, "/f"],
                                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    print(f"[server] killed PID {pid}")
                    break
        else:
            subprocess.call(["fuser", "-k", f"{port}/tcp"],
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as exc:
        print(f"[server] could not free port {port}: {exc}")
    time.sleep(0.6)   # brief wait for the OS to release the socket


if __name__ == "__main__":
    import uvicorn
    _PORT = 5000
    _free_port(_PORT)
    print("\n" + "="*60)
    print("  PTP Dashboard SSH Bridge — server running")
    print(f"  Open dashboard at:  http://127.0.0.1:{_PORT}/analysis.html")
    print("="*60 + "\n")
    uvicorn.run(app, host="127.0.0.1", port=_PORT, log_level="info")
